#!/usr/bin/env python3
"""
将指定列设为「文本」格式，避免纯数字（如长 ID、证件号）被显示为科学计数法。
用法:
  python format_columns_as_text.py --input data.xlsx --columns 身份证号,订单号 --output out.xlsx
  python format_columns_as_text.py --input data.xlsx --columns A,B,C --output out.xlsx
  python format_columns_as_text.py --input data.xlsx --columns 编号 --sheet 0 --output out.xlsx
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl


def _col_name_to_index(name):
    """列名（A, B, ..., Z, AA, ...）转 1-based 列索引。"""
    name = name.upper().strip()
    if not re.match(r"^[A-Z]+$", name):
        return None
    n = 0
    for c in name:
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n


def main():
    parser = argparse.ArgumentParser(description="将指定列设为文本格式，避免科学计数法")
    parser.add_argument("--input", "-i", required=True, help="输入 .xlsx 文件")
    parser.add_argument("--output", "-o", help="输出 .xlsx，默认覆盖原文件")
    parser.add_argument("--sheet", default=0, help="工作表名或索引")
    parser.add_argument("--columns", "-c", required=True, help="要设为文本的列，逗号分隔：列名（如 身份证号）或列字母（如 A,B）")
    parser.add_argument("--header-row", type=int, default=1, help="表头行号（1-based），该行也一并设为文本格式")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"文件不存在: {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    if isinstance(args.sheet, int):
        ws = wb.worksheets[args.sheet]
    else:
        ws = wb[args.sheet]

    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    header_row = args.header_row
    header = [ws.cell(row=header_row, column=j).value for j in range(1, max_col + 1)]

    col_specs = [s.strip() for s in args.columns.split(",") if s.strip()]
    col_indices = []
    for spec in col_specs:
        idx = _col_name_to_index(spec)
        if idx is not None:
            if 1 <= idx <= max_col:
                col_indices.append(idx)
            else:
                print(f"列字母越界: {spec}", file=sys.stderr)
                sys.exit(1)
        else:
            try:
                i = header.index(spec) + 1
                col_indices.append(i)
            except ValueError:
                print(f"表头中未找到列名: {spec}", file=sys.stderr)
                sys.exit(1)

    for col_idx in col_indices:
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.number_format = "@"
            val = cell.value
            if val is not None and not isinstance(val, str):
                cell.value = str(val)

    out = Path(args.output) if args.output else path
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已将 {len(col_indices)} 列设为文本格式 -> {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
