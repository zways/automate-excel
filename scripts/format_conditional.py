#!/usr/bin/env python3
"""
为 Excel 指定区域添加条件格式：大于/小于/等于/介于/重复值，或双色/三色色阶。
用法:
  python format_conditional.py --input data.xlsx --range "C2:C100" --rule gt --value 100 --fill "FF6B6B"
  python format_conditional.py --input data.xlsx --column D --rule duplicates --fill yellow
  python format_conditional.py --input data.xlsx --range "E2:E50" --rule between --min 0 --max 1 --fill "90EE90"
  python format_conditional.py --input data.xlsx --range "F2:F100" --color-scale
"""
import argparse
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, Rule
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle


# 常用颜色名 -> 六位十六进制（无前缀）
COLORS = {
    "red": "FF6B6B",
    "green": "90EE90",
    "yellow": "FFFF00",
    "orange": "FFA500",
    "blue": "87CEEB",
    "gray": "D3D3D3",
}


def _parse_fill(fill_arg):
    s = str(fill_arg).strip()
    if s.startswith("#"):
        s = s[1:]
    if len(s) == 6 and re.match(r"[0-9A-Fa-f]{6}", s):
        return s
    return COLORS.get(s.lower(), "FF6B6B")


def _resolve_range(ws, column_arg, range_arg):
    if range_arg:
        return range_arg
    if not column_arg:
        return None
    col = column_arg.upper()
    if len(col) == 1 and col.isalpha():
        max_row = ws.max_row or 1
        return f"{col}2:{col}{max(2, max_row)}"
    return None


def main():
    parser = argparse.ArgumentParser(description="为 Excel 区域添加条件格式")
    parser.add_argument("--input", "-i", required=True, help="输入 .xlsx 文件")
    parser.add_argument("--output", "-o", help="输出 .xlsx，默认覆盖原文件")
    parser.add_argument("--sheet", default=0, help="工作表名或索引")
    parser.add_argument("--range", "-r", help="区域，如 A2:A100（与 --column 二选一）")
    parser.add_argument("--column", "-c", help="列字母，如 A；将对该列从第 2 行到最大行应用格式")
    parser.add_argument("--rule", choices=("gt", "gte", "lt", "lte", "eq", "between", "duplicates"), help="条件：大于/大于等于/小于/小于等于/等于/介于/重复值")
    parser.add_argument("--value", help="--rule gt/lt/eq 等时的比较值")
    parser.add_argument("--min", help="--rule between 时的下限")
    parser.add_argument("--max", help="--rule between 时的上限")
    parser.add_argument("--fill", default="FF6B6B", help="满足条件时的填充色，如 red 或 FF6B6B")
    parser.add_argument("--color-scale", action="store_true", help="使用双色色阶（红-绿），忽略 --rule/--value")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"文件不存在: {path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    if isinstance(args.sheet, int):
        ws = wb.worksheets[args.sheet]
    else:
        ws = wb[args.sheet]

    cell_range = _resolve_range(ws, args.column, args.range)
    if not cell_range:
        print("请指定 --range 或 --column", file=sys.stderr)
        sys.exit(1)

    fill_hex = _parse_fill(args.fill)
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    dxf = DifferentialStyle(fill=fill)

    if args.color_scale:
        from openpyxl.formatting.rule import ColorScaleRule
        rule = ColorScaleRule(
            start_type="min", start_value=0, start_color="F8696B",
            end_type="max", end_value=100, end_color="63BE7B",
        )
        ws.conditional_formatting.add(cell_range, rule)
        print(f"已添加色阶条件格式: {cell_range}", file=sys.stderr)
    elif args.rule == "duplicates":
        rule = Rule(type="duplicateValues", dxf=dxf)
        ws.conditional_formatting.add(cell_range, rule)
        print(f"已添加重复值条件格式: {cell_range}", file=sys.stderr)
    elif args.rule == "between":
        if args.min is None or args.max is None:
            print("--rule between 需同时指定 --min 和 --max", file=sys.stderr)
            sys.exit(1)
        rule = CellIsRule(operator="between", formula=[str(args.min), str(args.max)], fill=fill)
        ws.conditional_formatting.add(cell_range, rule)
        print(f"已添加介于 {args.min}-{args.max} 条件格式: {cell_range}", file=sys.stderr)
    else:
        op_map = {"gt": "greaterThan", "gte": "greaterThanOrEqual", "lt": "lessThan", "lte": "lessThanOrEqual", "eq": "equal"}
        if args.rule not in op_map or not args.value:
            print("--rule gt/lt/eq 等需指定 --value", file=sys.stderr)
            sys.exit(1)
        rule = CellIsRule(operator=op_map[args.rule], formula=[str(args.value)], fill=fill)
        ws.conditional_formatting.add(cell_range, rule)
        print(f"已添加 {args.rule} {args.value} 条件格式: {cell_range}", file=sys.stderr)

    out = Path(args.output) if args.output else path
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已保存 -> {out}", file=sys.stderr)


if __name__ == "__main__":
    main()
