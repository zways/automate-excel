#!/usr/bin/env python3
"""
重命名 Excel 中的工作表。按原名或索引指定要重命名的 sheet。
用法:
  python rename_sheets.py --input file.xlsx --rename "Sheet1:汇总" "Sheet2:明细"
  python rename_sheets.py --input file.xlsx --rename "0:首页" "1:数据"
  python rename_sheets.py --input file.xlsx --prefix "2024_" --output out.xlsx
"""
import argparse
import sys
from pathlib import Path

import openpyxl


def main():
    parser = argparse.ArgumentParser(description="重命名 Excel 工作表")
    parser.add_argument("--input", "-i", required=True, help="输入 .xlsx 文件")
    parser.add_argument("--output", "-o", help="输出 .xlsx，默认覆盖原文件")
    parser.add_argument("--rename", "-r", action="append", help="原名:新名 或 索引:新名（索引从 0 起），可多次")
    parser.add_argument("--prefix", help="为所有 sheet 名前加前缀")
    parser.add_argument("--suffix", help="为所有 sheet 名后加后缀")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"文件不存在: {path}", file=sys.stderr)
        sys.exit(1)

    if not args.rename and not args.prefix and not args.suffix:
        print("请指定 --rename、--prefix 或 --suffix 之一", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    names = wb.sheetnames

    if args.rename:
        for item in args.rename:
            if ":" not in item:
                print(f"无效项（需 原名:新名 或 索引:新名）: {item}", file=sys.stderr)
                sys.exit(1)
            left, new_name = item.split(":", 1)
            left, new_name = left.strip(), new_name.strip()
            if not new_name:
                continue
            if left.isdigit():
                idx = int(left)
                if idx < 0 or idx >= len(names):
                    print(f"索引越界: {idx}", file=sys.stderr)
                    sys.exit(1)
                old_name = names[idx]
            else:
                old_name = left
                if old_name not in names:
                    print(f"工作表不存在: {old_name}", file=sys.stderr)
                    sys.exit(1)
            if len(new_name) > 31:
                new_name = new_name[:31]
            if new_name != old_name and new_name in names:
                print(f"新名已存在: {new_name}", file=sys.stderr)
                sys.exit(1)
            wb[old_name].title = new_name
            names = wb.sheetnames

    if args.prefix or args.suffix:
        for i, old in enumerate(wb.sheetnames):
            new = (args.prefix or "") + old + (args.suffix or "")
            if len(new) > 31:
                new = new[:31]
            if new != old:
                wb[old].title = new

    out = Path(args.output) if args.output else path
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已保存 -> {out}", file=sys.stderr)
    print(" ".join(wb.sheetnames))


if __name__ == "__main__":
    main()
