#!/usr/bin/env python3
"""
用数据表按行填充模板：模板中 {{列名}} 占位符被替换为数据行对应列的值，每行数据生成一行，输出到同一表。
用法:
  python template_fill.py --template template.xlsx --data data.xlsx --output out.xlsx
  python template_fill.py --template template.xlsx --data data.csv --pattern-row 2 --output out.xlsx
"""
import argparse
import re
import sys
from pathlib import Path

import pandas as pd
import openpyxl


def _replace_placeholders(text, row_dict):
    if text is None or not isinstance(text, str):
        return text
    result = text
    for col, val in row_dict.items():
        placeholder = "{{" + str(col) + "}}"
        if placeholder in result:
            result = result.replace(placeholder, str(val) if pd.notna(val) else "")
    # 未匹配的 {{xxx}} 保留或清空
    result = re.sub(r"\{\{[^}]+\}\}", "", result)
    return result


def main():
    parser = argparse.ArgumentParser(description="用数据表按行填充模板中的 {{列名}} 占位符")
    parser.add_argument("--template", "-t", required=True, help="模板 .xlsx，含占位符的一行作为样式来源")
    parser.add_argument("--data", "-d", required=True, help="数据 .xlsx 或 .csv")
    parser.add_argument("--output", "-o", required=True, help="输出 .xlsx")
    parser.add_argument("--sheet-template", default=0, help="模板 sheet 名或索引")
    parser.add_argument("--sheet-data", default=0, help="数据 sheet 名或索引（仅 .xlsx）")
    parser.add_argument("--pattern-row", type=int, default=2, help="模板中占位符所在行号（1-based），默认第 2 行")
    parser.add_argument("--header-row", type=int, default=1, help="模板表头行号（1-based），默认第 1 行")
    args = parser.parse_args()

    tpath = Path(args.template)
    dpath = Path(args.data)
    if not tpath.exists():
        print(f"模板不存在: {tpath}", file=sys.stderr)
        sys.exit(1)
    if not dpath.exists():
        print(f"数据文件不存在: {dpath}", file=sys.stderr)
        sys.exit(1)

    if dpath.suffix.lower() == ".csv":
        data_df = pd.read_csv(dpath, encoding="utf-8-sig")
    else:
        data_df = pd.read_excel(dpath, sheet_name=args.sheet_data, engine="openpyxl")

    wb = openpyxl.load_workbook(tpath)
    if isinstance(args.sheet_template, int):
        ws = wb.worksheets[args.sheet_template]
    else:
        ws = wb[args.sheet_template]

    pattern_row_idx = args.pattern_row
    max_col = ws.max_column

    for i, row in enumerate(data_df.itertuples(index=False)):
        row_dict = dict(zip(data_df.columns, row))
        target_row = pattern_row_idx + i
        for col_idx in range(1, max_col + 1):
            src = ws.cell(row=pattern_row_idx, column=col_idx)
            cell = ws.cell(row=target_row, column=col_idx)
            cell.value = _replace_placeholders(src.value, row_dict)
            try:
                if src.has_style and src.font:
                    cell.font = src.font.copy()
                if src.has_style and src.border:
                    cell.border = src.border.copy()
                if src.has_style and src.fill:
                    cell.fill = src.fill.copy()
                if src.number_format:
                    cell.number_format = src.number_format
                if src.has_style and src.alignment:
                    cell.alignment = src.alignment.copy()
            except Exception:
                pass

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已用 {len(data_df)} 行数据填充模板 -> {out}")


if __name__ == "__main__":
    main()
